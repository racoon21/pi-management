import re
from dataclasses import dataclass, field
from io import BytesIO
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Task, TaskHistory
from app.services.task_service import _task_to_snapshot
from app.schemas.upload import (
    ExcelRow,
    HierarchyNode,
    UploadPreview,
    DiffNode,
    DiffResult,
    UpsertResult,
)

VALID_ORG_TYPES = ["본부", "실", "담당", "팀"]

# 한글 헤더 → 필드명 매핑
EXTRA_COL_MAP = {
    "조직단위": "organization_type",
    "조직명": "organization_name",
    "담당자": "manager_name",
    "사번": "manager_id",
    "키워드": "keywords",
    "AI활용": "is_ai_utilized",
}


@dataclass
class ParsedExcel:
    rows: list[ExcelRow] = field(default_factory=list)


@dataclass
class L4Entry:
    name: str
    related_team: str = ""
    organization_type: str = ""
    organization_name: str = ""
    manager_name: str = ""
    manager_id: str = ""
    keywords: str = ""
    is_ai_utilized: str = ""


def parse_excel(file_bytes: bytes) -> ParsedExcel:
    """openpyxl로 엑셀 파싱. 헤더에서 L1~L4 + 추가 컬럼 자동 감지."""
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # 헤더 행에서 컬럼 인덱스 찾기
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map: dict[str, int] = {}
    related_team_cols: dict[str, int] = {}
    extra_cols: dict[str, int] = {}  # field_name -> column index

    for idx, cell_value in enumerate(header_row):
        if cell_value is None:
            continue
        val = str(cell_value).strip().upper()
        raw = str(cell_value).strip()

        if val in ("L1", "L2", "L3", "L4"):
            col_map[val] = idx
        # 유관팀 컬럼 감지
        if "유관팀" in raw:
            if "L3" in raw.upper():
                related_team_cols["L3"] = idx
            elif "L4" in raw.upper():
                related_team_cols["L4"] = idx
            else:
                if "L3" in col_map and idx == col_map["L3"] + 1:
                    related_team_cols["L3"] = idx
                elif "L4" in col_map and idx == col_map["L4"] + 1:
                    related_team_cols["L4"] = idx
        # 추가 컬럼 감지
        if raw in EXTRA_COL_MAP:
            extra_cols[EXTRA_COL_MAP[raw]] = idx

    if not all(k in col_map for k in ("L1", "L2", "L3", "L4")):
        raise ValueError(
            f"엑셀 헤더에서 L1~L4 컬럼을 찾을 수 없습니다. 발견된 컬럼: {list(col_map.keys())}"
        )

    def _cell(row_data: tuple, idx: int) -> str:
        if idx < len(row_data) and row_data[idx] is not None:
            return str(row_data[idx]).strip()
        return ""

    # Forward-fill 변수
    prev_l1 = prev_l2 = prev_l3 = ""

    rows: list[ExcelRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        l1 = _cell(row, col_map["L1"])
        l2 = _cell(row, col_map["L2"])
        l3 = _cell(row, col_map["L3"])
        l4 = _cell(row, col_map["L4"])

        # Forward-fill: 빈 셀은 이전 값 계승
        if l1:
            prev_l1 = l1
        else:
            l1 = prev_l1
        if l2:
            prev_l2 = l2
        else:
            l2 = prev_l2
        if l3:
            prev_l3 = l3
        else:
            l3 = prev_l3

        # 빈 행 건너뛰기 (L4가 없으면 유효하지 않은 행)
        if not l4:
            continue

        l3_rt = _cell(row, related_team_cols["L3"]) if "L3" in related_team_cols else ""
        l4_rt = _cell(row, related_team_cols["L4"]) if "L4" in related_team_cols else ""

        # 추가 컬럼 읽기
        extra_vals: dict[str, str] = {}
        for field_name, col_idx in extra_cols.items():
            extra_vals[field_name] = _cell(row, col_idx)

        rows.append(
            ExcelRow(
                l1=l1,
                l2=l2,
                l3=l3,
                l4=l4,
                l3_related_team=l3_rt,
                l4_related_team=l4_rt,
                organization_type=extra_vals.get("organization_type", ""),
                organization_name=extra_vals.get("organization_name", ""),
                manager_name=extra_vals.get("manager_name", ""),
                manager_id=extra_vals.get("manager_id", ""),
                keywords=extra_vals.get("keywords", ""),
                is_ai_utilized=extra_vals.get("is_ai_utilized", ""),
            )
        )

    wb.close()

    # 유효성 검증
    for i, row in enumerate(rows):
        if row.organization_type and row.organization_type not in VALID_ORG_TYPES:
            raise ValueError(
                f"행 {i + 2}: 조직단위 '{row.organization_type}'는 유효하지 않습니다. "
                f"허용값: {VALID_ORG_TYPES}"
            )
        if row.is_ai_utilized:
            upper = row.is_ai_utilized.upper()
            if upper not in ("Y", "N", "YES", "NO", "TRUE", "FALSE"):
                raise ValueError(
                    f"행 {i + 2}: AI활용 '{row.is_ai_utilized}'는 유효하지 않습니다. "
                    f"허용값: Y, N"
                )

    return ParsedExcel(rows=rows)


def _parse_related_team(value: str) -> list[str] | None:
    """Parse comma-separated related_team string into a list."""
    if not value:
        return None
    teams = [t.strip() for t in value.split(",") if t.strip()]
    return teams if teams else None


def _parse_keywords(value: str) -> list[str] | None:
    """Parse comma-separated keywords string into a list."""
    if not value:
        return None
    items = [k.strip() for k in value.split(",") if k.strip()]
    return items if items else None


def _parse_ai_utilized(value: str) -> bool:
    """Parse Y/N/Yes/No/TRUE/FALSE to bool."""
    if not value:
        return False
    return value.strip().upper() in ("Y", "YES", "TRUE")


def build_hierarchy(parsed: ParsedExcel) -> list[HierarchyNode]:
    """파싱된 데이터를 계층 트리로 변환."""
    # l1 -> { l2 -> { l3 -> [L4Entry] } }
    tree: dict[str, dict] = {}
    l3_rt_map: dict[tuple[str, str, str], str] = {}

    for row in parsed.rows:
        if row.l1 not in tree:
            tree[row.l1] = {}
        l2_map = tree[row.l1]

        if row.l2 not in l2_map:
            l2_map[row.l2] = {}
        l3_map = l2_map[row.l2]

        if row.l3 not in l3_map:
            l3_map[row.l3] = []
        l4_list: list[L4Entry] = l3_map[row.l3]

        # Store L3 related_team (first non-empty wins)
        l3_key = (row.l1, row.l2, row.l3)
        if l3_key not in l3_rt_map and row.l3_related_team:
            l3_rt_map[l3_key] = row.l3_related_team

        # 중복 L4 방지
        if not any(item.name == row.l4 for item in l4_list):
            l4_list.append(
                L4Entry(
                    name=row.l4,
                    related_team=row.l4_related_team,
                    organization_type=row.organization_type,
                    organization_name=row.organization_name,
                    manager_name=row.manager_name,
                    manager_id=row.manager_id,
                    keywords=row.keywords,
                    is_ai_utilized=row.is_ai_utilized,
                )
            )

    # dict → HierarchyNode 트리
    result: list[HierarchyNode] = []
    for l1_name, l2_map in tree.items():
        l2_nodes: list[HierarchyNode] = []
        for l2_name, l3_map in l2_map.items():
            l3_nodes: list[HierarchyNode] = []
            for l3_name, l4_list in l3_map.items():
                l4_nodes = [
                    HierarchyNode(
                        name=entry.name,
                        level="L4",
                        related_team=_parse_related_team(entry.related_team),
                        organization_type=entry.organization_type or None,
                        organization_name=entry.organization_name or None,
                        manager_name=entry.manager_name or None,
                        manager_id=entry.manager_id or None,
                        keywords=_parse_keywords(entry.keywords),
                        is_ai_utilized=_parse_ai_utilized(entry.is_ai_utilized),
                    )
                    for entry in l4_list
                ]
                l3_rt = l3_rt_map.get((l1_name, l2_name, l3_name), "")
                l3_nodes.append(HierarchyNode(
                    name=l3_name, level="L3",
                    related_team=_parse_related_team(l3_rt),
                    children=l4_nodes,
                ))
            l2_nodes.append(HierarchyNode(name=l2_name, level="L2", children=l3_nodes))
        result.append(HierarchyNode(name=l1_name, level="L1", children=l2_nodes))

    return result


def build_preview(parsed: ParsedExcel) -> UploadPreview:
    """미리보기 데이터 생성."""
    unique_l1 = set()
    unique_l2 = set()
    unique_l3 = set()
    unique_l4 = set()

    for row in parsed.rows:
        unique_l1.add(row.l1)
        unique_l2.add((row.l1, row.l2))
        unique_l3.add((row.l1, row.l2, row.l3))
        unique_l4.add((row.l1, row.l2, row.l3, row.l4))

    return UploadPreview(
        rows=parsed.rows[:10],
        total_rows=len(parsed.rows),
        summary={
            "l1_count": len(unique_l1),
            "l2_count": len(unique_l2),
            "l3_count": len(unique_l3),
            "l4_count": len(unique_l4),
        },
        hierarchy=build_hierarchy(parsed),
    )


async def diff_tasks(db: AsyncSession, parsed: ParsedExcel) -> DiffResult:
    """파싱된 데이터를 기존 DB와 비교하여 diff 트리 반환."""
    result = await db.execute(
        select(Task).where(Task.level == "Root", Task.deleted_at.is_(None))
    )
    root = result.scalar_one_or_none()

    all_tasks_result = await db.execute(
        select(Task).where(Task.deleted_at.is_(None))
    )
    all_tasks = list(all_tasks_result.scalars().all())

    task_by_parent_name: dict[tuple[UUID | None, str], Task] = {}
    for t in all_tasks:
        normalized = _normalize_name(t.name)
        task_by_parent_name[(t.parent_id, normalized)] = t

    hierarchy = build_hierarchy(parsed)
    stats = {"new": 0, "existing": 0, "total": 0}

    def diff_node(node: HierarchyNode, parent_id: UUID | None) -> DiffNode:
        existing = task_by_parent_name.get((parent_id, _normalize_name(node.name)))
        status = "existing" if existing else "new"
        stats[status] += 1
        stats["total"] += 1

        children: list[DiffNode] = []
        child_parent_id = existing.id if existing else None
        for child in node.children:
            children.append(diff_node(child, child_parent_id))

        return DiffNode(
            name=node.name,
            level=node.level,
            status=status,
            children=children,
        )

    diff_tree: list[DiffNode] = []
    root_id = root.id if root else None
    for l1_node in hierarchy:
        diff_tree.append(diff_node(l1_node, root_id))

    return DiffResult(diff_tree=diff_tree, stats=stats)


async def upsert_tasks(
    db: AsyncSession, parsed: ParsedExcel, user_id: UUID
) -> UpsertResult:
    """파싱된 데이터를 DB에 upsert (Bulk 최적화)."""
    created = 0
    skipped = 0

    # 기존 태스크 전체 로드 + 인메모리 인덱스
    all_result = await db.execute(
        select(Task).where(Task.deleted_at.is_(None))
    )
    all_tasks = list(all_result.scalars().all())

    index: dict[tuple[UUID | None, str], Task] = {}
    for t in all_tasks:
        index[(t.parent_id, _normalize_name(t.name))] = t

    # Root 노드 조회/생성
    root = next((t for t in all_tasks if t.level == "Root"), None)
    if not root:
        root = Task(
            level="Root", name="Root", organization="",
            created_by=user_id, updated_by=user_id,
        )
        db.add(root)
        await db.flush()
        _create_history(db, root, user_id)
        created += 1

    hierarchy = build_hierarchy(parsed)

    new_tasks: list[Task] = []

    def _lookup_or_new(
        parent: Task, level: str, name: str, organization: str,
        related_team: list[str] | None = None,
        organization_type: str | None = None,
        organization_name: str | None = None,
        manager_name: str | None = None,
        manager_id: str | None = None,
        keywords: list[str] | None = None,
        is_ai_utilized: bool = False,
    ) -> Task:
        nonlocal created, skipped
        key = (parent.id, _normalize_name(name))
        existing = index.get(key)
        if existing:
            skipped += 1
            return existing
        task = Task(
            parent_id=parent.id, level=level, name=name,
            organization=organization,
            organization_type=organization_type,
            organization_name=organization_name,
            manager_name=manager_name,
            manager_id=manager_id,
            related_team=related_team,
            keywords=keywords or [],
            is_ai_utilized=is_ai_utilized,
            created_by=user_id, updated_by=user_id,
        )
        new_tasks.append(task)
        created += 1
        return task

    # L1 수집
    l1_map: list[tuple[Task, HierarchyNode]] = []
    for l1_node in hierarchy:
        l1_task = _lookup_or_new(root, "L1", l1_node.name, l1_node.name)
        l1_map.append((l1_task, l1_node))

    await _flush_new(db, new_tasks, index)

    # L2 수집
    l2_map: list[tuple[Task, HierarchyNode, str]] = []
    for l1_task, l1_node in l1_map:
        for l2_node in l1_node.children:
            l2_task = _lookup_or_new(l1_task, "L2", l2_node.name, l1_node.name)
            l2_map.append((l2_task, l2_node, l1_node.name))

    await _flush_new(db, new_tasks, index)

    # L3 수집
    l3_map: list[tuple[Task, HierarchyNode, str]] = []
    for l2_task, l2_node, org in l2_map:
        for l3_node in l2_node.children:
            l3_task = _lookup_or_new(
                l2_task, "L3", l3_node.name, org,
                related_team=l3_node.related_team,
            )
            l3_map.append((l3_task, l3_node, org))

    await _flush_new(db, new_tasks, index)

    # L4 수집 — 추가 필드 전달
    for l3_task, l3_node, org in l3_map:
        for l4_node in l3_node.children:
            _lookup_or_new(
                l3_task, "L4", l4_node.name, org,
                related_team=l4_node.related_team,
                organization_type=l4_node.organization_type,
                organization_name=l4_node.organization_name,
                manager_name=l4_node.manager_name,
                manager_id=l4_node.manager_id,
                keywords=l4_node.keywords,
                is_ai_utilized=l4_node.is_ai_utilized,
            )

    await _flush_new(db, new_tasks, index)

    await db.commit()
    return UpsertResult(created=created, skipped=skipped, total=created + skipped)


async def _flush_new(
    db: AsyncSession,
    pending: list[Task],
    index: dict[tuple[UUID | None, str], Task],
) -> None:
    """pending 리스트의 신규 태스크를 bulk flush + 히스토리 생성 + 인덱스 갱신."""
    if not pending:
        return
    batch = list(pending)
    pending.clear()
    db.add_all(batch)
    await db.flush()
    for task in batch:
        _create_history(db, task, task.created_by)
        index[(task.parent_id, _normalize_name(task.name))] = task


def _create_history(db: AsyncSession, task: Task, user_id: UUID) -> None:
    """태스크 생성 히스토리 기록."""
    history = TaskHistory(
        task_id=task.id,
        snapshot=_task_to_snapshot(task),
        version=1,
        change_type="CREATE",
        changed_by=user_id,
    )
    db.add(history)


def _normalize_name(name: str) -> str:
    """[IMP-08] 비교용 정규화: 모든 공백 제거"""
    return re.sub(r"\s+", "", name.strip())
