from io import BytesIO

from fastapi import APIRouter, HTTPException, UploadFile, File, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.api.deps import DbSession, EditorUser
from app.schemas.common import ApiResponse
from app.schemas.upload import UploadPreview, DiffResult, UpsertResult
from app.services import upload_service
from app.core.cache import task_cache

router = APIRouter(prefix="/upload", tags=["upload"])

ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB
CHUNK_SIZE = 1024 * 1024  # 1MB


async def _read_file_chunked(file: UploadFile) -> bytes:
    """Read uploaded file in 1MB chunks, enforcing MAX_FILE_SIZE."""
    buf = BytesIO()
    total = 0
    while True:
        chunk = await file.read(CHUNK_SIZE)
        if not chunk:
            break
        total += len(chunk)
        if total > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail="파일 크기가 100MB를 초과합니다.",
            )
        buf.write(chunk)
    return buf.getvalue()


def _validate_file(file: UploadFile) -> None:
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="파일명이 없습니다.",
        )
    ext = "." + file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"지원하지 않는 파일 형식입니다. (.xlsx, .xls만 허용)",
        )


@router.post("/preview", response_model=ApiResponse[UploadPreview])
async def upload_preview(
    current_user: EditorUser,
    file: UploadFile = File(...),
):
    """엑셀 파일을 파싱하여 미리보기 데이터를 반환합니다."""
    _validate_file(file)
    file_bytes = await _read_file_chunked(file)

    try:
        parsed = upload_service.parse_excel(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="엑셀 파일을 파싱할 수 없습니다. 올바른 형식인지 확인해주세요.",
        )

    if not parsed.rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="엑셀 파일에 유효한 데이터가 없습니다.",
        )

    preview = upload_service.build_preview(parsed)
    return ApiResponse(success=True, data=preview)


@router.post("/diff", response_model=ApiResponse[DiffResult])
async def upload_diff(
    db: DbSession,
    current_user: EditorUser,
    file: UploadFile = File(...),
):
    """엑셀 파일을 파싱하여 기존 DB와 비교한 diff를 반환합니다."""
    _validate_file(file)
    file_bytes = await _read_file_chunked(file)

    try:
        parsed = upload_service.parse_excel(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="엑셀 파일을 파싱할 수 없습니다.",
        )

    diff = await upload_service.diff_tasks(db, parsed)
    return ApiResponse(success=True, data=diff)


@router.post("/confirm", response_model=ApiResponse[UpsertResult])
async def upload_confirm(
    db: DbSession,
    current_user: EditorUser,
    file: UploadFile = File(...),
):
    """엑셀 파일을 파싱하여 DB에 upsert합니다."""
    _validate_file(file)
    file_bytes = await _read_file_chunked(file)

    try:
        parsed = upload_service.parse_excel(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="엑셀 파일을 파싱할 수 없습니다.",
        )

    result = await upload_service.upsert_tasks(db, parsed, current_user.id)
    task_cache.invalidate()
    return ApiResponse(success=True, data=result)


@router.get("/template")
async def download_template():
    """엑셀 업로드 양식 다운로드."""
    wb = Workbook()
    ws = wb.active
    ws.title = "업무PI양식"

    headers = [
        "L1", "L2", "L3", "L3 유관팀", "L4", "L4 유관팀",
        "조직단위", "조직명", "담당자", "사번", "키워드", "AI활용",
    ]
    col_widths = [20, 25, 25, 20, 40, 20, 12, 15, 12, 12, 25, 10]

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7952B3", end_color="7952B3", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    # 샘플 데이터 행
    sample_row = [
        "유선사업본부", "통합 마케팅전략 수립", "유선사업전략 수립/실행",
        "AI보드, Infra운용팀",
        "(분석) 시장, 경쟁 Trend 및 고객 Data 기반 전략 방향성 수립", "",
        "본부", "마케팅팀", "홍길동", "A12345", "마케팅, 전략, 분석", "N",
    ]
    sample_font = Font(color="999999", italic=True)
    for col_idx, value in enumerate(sample_row, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = sample_font
        cell.border = thin_border

    # 데이터 유효성 검사: 조직단위 (G열, 2~1000행)
    org_type_dv = DataValidation(
        type="list",
        formula1='"본부,실,담당,팀"',
        allow_blank=True,
    )
    org_type_dv.error = "본부, 실, 담당, 팀 중 선택해주세요"
    org_type_dv.errorTitle = "유효하지 않은 조직단위"
    ws.add_data_validation(org_type_dv)
    org_type_dv.add("G2:G1000")

    # 데이터 유효성 검사: AI활용 (L열, 2~1000행)
    ai_dv = DataValidation(
        type="list",
        formula1='"Y,N"',
        allow_blank=True,
    )
    ai_dv.error = "Y 또는 N을 선택해주세요"
    ai_dv.errorTitle = "유효하지 않은 AI활용 값"
    ws.add_data_validation(ai_dv)
    ai_dv.add("L2:L1000")

    # 헤더 행 고정
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename = "조직_업무PI_양식_ver0.9.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}",
        },
    )
